"""
Report Generator for Factory Process Monitoring Agent
Generates PDF and Excel reports for production data
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportGenerator:
    """Generate PDF and Excel reports for factory monitoring data"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a2347'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=12,
            spaceBefore=12
        )
    
    def generate_pdf_report(self, production_lines, quality_metrics, overall_metrics, alerts):
        """Generate a comprehensive PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Title
        title = Paragraph("Factory Process Monitoring Report", self.title_style)
        elements.append(title)
        
        # Report metadata
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        metadata = Paragraph(f"<b>Generated:</b> {timestamp}", self.styles['Normal'])
        elements.append(metadata)
        elements.append(Spacer(1, 20))
        
        # Overall Metrics Section
        elements.append(Paragraph("Overall Production Metrics", self.heading_style))
        
        overall_data = [
            ['Metric', 'Value'],
            ['Total Output', f"{overall_metrics['total_output']:,} units"],
            ['Total Defects', f"{overall_metrics['total_defects']:,} units"],
            ['Overall OEE', f"{overall_metrics['overall_oee']}%"],
            ['Average Efficiency', f"{overall_metrics['average_efficiency']}%"],
            ['Active Lines', f"{overall_metrics['active_lines']}/{overall_metrics['total_lines']}"],
            ['Critical Alerts', str(overall_metrics['critical_alerts'])],
            ['Warning Alerts', str(overall_metrics['warning_alerts'])]
        ]
        
        overall_table = Table(overall_data, colWidths=[3*inch, 3*inch])
        overall_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(overall_table)
        elements.append(Spacer(1, 20))
        
        # Production Lines Section
        elements.append(Paragraph("Production Lines Status", self.heading_style))
        
        line_data = [['Line ID', 'Status', 'Speed', 'Efficiency', 'Output', 'Defects']]
        for line in production_lines:
            line_data.append([
                line['line_id'],
                line['status'].upper(),
                f"{line['current_speed']:.0f} u/min",
                f"{line['efficiency']:.1f}%",
                f"{line['products_produced']:,}",
                str(line['defects'])
            ])
        
        line_table = Table(line_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 1.2*inch, 0.8*inch])
        line_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 20))
        
        # Quality Metrics Section
        elements.append(Paragraph("Quality Control Metrics", self.heading_style))
        
        quality_data = [['Line ID', 'Inspected', 'Passed', 'Failed', 'Defect Rate', 'Quality Score']]
        for qm in quality_metrics:
            quality_data.append([
                qm['line_id'],
                f"{qm['total_inspected']:,}",
                f"{qm['passed']:,}",
                f"{qm['failed']:,}",
                f"{qm['defect_rate']:.2f}%",
                f"{qm['average_quality_score']:.1f}"
            ])
        
        quality_table = Table(quality_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 1*inch, 1.2*inch, 1.2*inch])
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(quality_table)
        elements.append(Spacer(1, 20))
        
        # Active Alerts Section
        if alerts:
            elements.append(Paragraph("Active Alerts", self.heading_style))
            
            alert_data = [['Severity', 'Line', 'Title', 'Message']]
            for alert in alerts[:10]:  # Limit to 10 most recent alerts
                alert_data.append([
                    alert['severity'].upper(),
                    alert['line_id'],
                    alert['title'],
                    alert['message'][:50] + '...' if len(alert['message']) > 50 else alert['message']
                ])
            
            alert_table = Table(alert_data, colWidths=[1*inch, 1*inch, 1.5*inch, 3.2*inch])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(alert_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_excel_report(self, production_lines, quality_metrics, overall_metrics, alerts):
        """Generate a comprehensive Excel report"""
        buffer = io.BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Header styles
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Overall Metrics Sheet
        ws_overall = wb.create_sheet("Overall Metrics")
        ws_overall.append(["Factory Process Monitoring Report"])
        ws_overall.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_overall.append([])
        ws_overall.append(["Metric", "Value"])
        
        overall_data = [
            ["Total Output", f"{overall_metrics['total_output']:,} units"],
            ["Total Defects", f"{overall_metrics['total_defects']:,} units"],
            ["Overall OEE", f"{overall_metrics['overall_oee']}%"],
            ["Average Efficiency", f"{overall_metrics['average_efficiency']}%"],
            ["Active Lines", f"{overall_metrics['active_lines']}/{overall_metrics['total_lines']}"],
            ["Critical Alerts", overall_metrics['critical_alerts']],
            ["Warning Alerts", overall_metrics['warning_alerts']]
        ]
        
        for row in overall_data:
            ws_overall.append(row)
        
        # Style overall metrics
        for cell in ws_overall[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        ws_overall.column_dimensions['A'].width = 25
        ws_overall.column_dimensions['B'].width = 20
        
        # Production Lines Sheet
        ws_lines = wb.create_sheet("Production Lines")
        line_headers = ["Line ID", "Name", "Status", "Current Speed", "Target Speed", 
                       "Efficiency (%)", "Uptime (%)", "Products Produced", "Defects"]
        ws_lines.append(line_headers)
        
        for line in production_lines:
            ws_lines.append([
                line['line_id'],
                line['name'],
                line['status'].upper(),
                f"{line['current_speed']:.2f}",
                f"{line['target_speed']:.2f}",
                f"{line['efficiency']:.2f}",
                f"{line['uptime']:.2f}",
                line['products_produced'],
                line['defects']
            ])
        
        # Style production lines
        for cell in ws_lines[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws_lines.column_dimensions[col].width = 15
        
        # Quality Metrics Sheet
        ws_quality = wb.create_sheet("Quality Metrics")
        quality_headers = ["Line ID", "Total Inspected", "Passed", "Failed", 
                          "Defect Rate (%)", "Quality Score", "Trend"]
        ws_quality.append(quality_headers)
        
        for qm in quality_metrics:
            ws_quality.append([
                qm['line_id'],
                qm['total_inspected'],
                qm['passed'],
                qm['failed'],
                f"{qm['defect_rate']:.2f}",
                f"{qm['average_quality_score']:.2f}",
                qm['trend']
            ])
        
        # Style quality metrics
        for cell in ws_quality[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_quality.column_dimensions[col].width = 18
        
        # Alerts Sheet
        if alerts:
            ws_alerts = wb.create_sheet("Active Alerts")
            alert_headers = ["Alert ID", "Line ID", "Severity", "Title", "Message", "Timestamp"]
            ws_alerts.append(alert_headers)
            
            for alert in alerts:
                ws_alerts.append([
                    alert['alert_id'],
                    alert['line_id'],
                    alert['severity'].upper(),
                    alert['title'],
                    alert['message'],
                    alert['timestamp']
                ])
            
            # Style alerts
            for cell in ws_alerts[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws_alerts.column_dimensions[col].width = 20
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer
